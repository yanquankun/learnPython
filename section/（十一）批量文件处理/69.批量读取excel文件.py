"""
@Desc: 本章节主要讲解如何批量读取 Excel 文件
@Author: Mint.Yan
@Date: 2025-07-211 15:04:12
"""

from openpyxl import Workbook, load_workbook
from pathlib import PurePath, Path
import os

excels_path = os.path.abspath('../../files/excels/')
os.chdir(excels_path)

p = Path('.')
# 只获取后缀为 .xlsx 的文件，并且文件名是纯数字
xlsx_files = [x for x in p.iterdir() if PurePath(x).suffix == '.xlsx' and x.stem.isdigit()]


# excel 的批量合并，我们以xlsx_files列表文件为目的，将其合并为一个excel文件

def merge_excel_files(files):
    merged_workbook = Workbook()
    merged_sheet = merged_workbook.active
    merged_sheet.title = "Merged Data"

    for file in files:
        workbook = load_workbook(file)
        sheet = workbook.active

        for row in sheet.iter_rows(values_only=True):
            merged_sheet.append(row)

    return merged_workbook


wb = merge_excel_files(xlsx_files)
wb.save('merged.xlsx')

print("===" * 10)

# 拆分 excel 文件

# 我们以 merged.xlsx 文件为目标，拆分为三个文件，每个文件为merged.xlsx的一列
# 分别为 merged_1.xlsx、merged_2.xlsx、merged_3.xlsx
# 最后输出 new_1.xlsx、new_2.xlsx、new_3.xlsx

wb = load_workbook('merged.xlsx')

# 激活工作表 Merged Data
wb.active = wb['Merged Data']
ws = wb.active

# 获取工作表
idx = 1
for col in ws.iter_cols():
    # 创建新的工作簿
    new_wb = Workbook()
    ws = new_wb.active
    ws.title = f"new sheet {idx}"
    for cell in col[:6]:  # 只取前6行数据
        if cell.value is not None:  # 检查单元格是否有值
            ws.append([cell.value])
    new_wb.save(f'new_{idx}.xlsx')
    print(f'new_{idx}.xlsx created successfully.')
    idx += 1
